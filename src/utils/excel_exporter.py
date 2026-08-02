"""Excel exporter - Génération de fichiers Excel structurés."""
import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelExporter:
    """Génère des fichiers Excel avec plusieurs onglets et mise en forme."""

    def __init__(self, filename: str = None, output_dir: Optional[str] = None):
        self.filename = filename or f"Mesures_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        self.output_dir = output_dir or os.getcwd()
        self.workbook = Workbook()

        # Styles
        self.style_header = self._create_style(
            font=Font(bold=True, size=12, color="FFFFFF"),
            fill=PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
            alignment=Alignment(horizontal="center")
        )

        self.style_alternate = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")

    def _create_style(self, font=None, fill=None, alignment=None):
        """Crée un style complet avec bordures."""
        return {
            'font': font,
            'fill': fill,
            'alignment': alignment,
            'borders': _THIN_BORDER,
        }

    def _apply_cell_style(self, cell, style: dict):
        """Applique un dictionnaire de style à une cellule."""
        if style.get('font'):
            cell.font = style['font']
        if style.get('fill'):
            cell.fill = style['fill']
        if style.get('alignment'):
            cell.alignment = style['alignment']
        if style.get('borders'):
            cell.border = style['borders']

    def _write_header_row(self, ws, headers, style: dict = None):
        """Écrit une ligne d'en-tête avec style."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            if style:
                self._apply_cell_style(cell, style)
            else:
                cell.border = _THIN_BORDER

    def _write_data_rows(self, ws, data_rows, start_row=2):
        """Écrit des lignes de données avec bordures."""
        for row_idx, row_data in enumerate(data_rows, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = _THIN_BORDER

    def add_summary_sheet(self, user_name: str, total_measures: int = 0):
        """Ajoute l'onglet Résumé."""
        ws = self.workbook.active

        # Titre principal
        ws.merge_cells('A1:D1')
        title_style = self._create_style(
            font=Font(bold=True, size=24, color="FFFFFF"),
            fill=PatternFill(start_color="3D3D50", end_color="3D3D50", fill_type="solid"),
            alignment=Alignment(horizontal="center")
        )
        ws['A1'] = "Résumé des Mesures"
        self._apply_cell_style(ws['A1'], title_style)
        # Appliquer aussi aux cellules merged pour le fond
        for cell_ref in ['B1', 'C1', 'D1']:
            self._apply_cell_style(ws[cell_ref], title_style)

        # Informations détaillées
        info_data = [
            ["Généré le:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Utilisateur:", user_name],
            ["Total des mesures:", str(total_measures)]
        ]

        for row_idx, row_data in enumerate(info_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = _THIN_BORDER
                if col_idx == 1:
                    cell.font = Font(bold=True)

    def add_all_measures_sheet(self, measurements: List[Dict]):
        """Ajoute l'onglet 'Toutes les mesures'."""
        ws = self.workbook.create_sheet("Toutes les mesures")
        headers = ["Horodatage", "Outil", "Type d'outil", "Valeur mesurée", "Unité", "Statut", "Notes"]
        self._write_header_row(ws, headers, self.style_header)

        data_rows = []
        for m in measurements:
            data_rows.append([
                m.get('timestamp', ''),
                m.get('tool_name', ''),
                m.get('tool_type', ''),
                m.get('value', ''),
                m.get('unit', ''),
                m.get('status', ''),
                m.get('note', ''),
            ])
        self._write_data_rows(ws, data_rows)

    def add_tool_sheets(self, measurements: List[Dict]):
        """Ajoute un onglet par type d'outil."""
        tool_types = {}
        for m in measurements:
            tool_type = m.get('tool_type', 'Other')
            if tool_type not in tool_types:
                tool_types[tool_type] = []
            tool_types[tool_type].append(m)

        for tool_type, tool_measurements in tool_types.items():
            sheet_name = tool_type[:31]
            counter = 1
            base_name = sheet_name
            while sheet_name in self.workbook.sheetnames:
                counter += 1
                suffix = f"_{counter}"
                sheet_name = f"{base_name[:31 - len(suffix)]}{suffix}"

            ws = self.workbook.create_sheet(sheet_name)
            headers = ["Horodatage", "Outil", "Valeur mesurée", "Unité", "Statut", "Notes"]
            self._write_header_row(ws, headers, self.style_header)

            data_rows = []
            for m in tool_measurements:
                data_rows.append([
                    m.get('timestamp', ''),
                    m.get('tool_name', ''),
                    m.get('value', ''),
                    m.get('unit', ''),
                    m.get('status', ''),
                    m.get('note', ''),
                ])
            self._write_data_rows(ws, data_rows)

    def save(self) -> str:
        """Sauvegarde le fichier Excel et retourne son chemin."""
        file_path = os.path.join(self.output_dir, self.filename)
        self.workbook.save(file_path)
        return file_path

    def get_filename(self) -> str:
        """Retourne le nom du fichier généré."""
        return self.filename
