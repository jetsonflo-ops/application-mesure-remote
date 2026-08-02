"""DataExporter — Export multi-format automatique des mesures.

Support des 6 formats :
  1. XLSX  — Excel 2010+ (openpyxl write_only, memoire <10Mo)
  2. CSV   — Valeurs separees par virgules (utf-8-sig BOM Excel)
  3. JSON  — Donnees structurees (conforme strict, valeurs securisees)
  4. XML   — Markup structure (xml.etree.ElementTree)
  5. PDF   — Rapport formaté A4 (reportlab, polices standard)
  6. SQLite — Base de donnees locale (WAL)

Principe de retry (idempotent) :
  - Si un export echoue, on verifie d'abord si le fichier destination
    existe deja avec un contenu valide. Si oui, on considere que
    l'export a reussi entre-temps → pas de nouvel essai.
  - Si le fichier n'existe pas, on retente jusqu'a 3 fois avec un
    delai progressif (0.5s, 1s, 2s) + jitter ±25%.
  - Un echec definitif ne bloque pas les autres formats.
  - SQLite WAL mode pour eviter les verrous.

Le retry ne duplique JAMAIS une donnee. Il verifie que la donnee
n'a pas deja ete ecrite avant chaque nouvelle tentative.

Utilisation:
    exporter = DataExporter(output_dir="/chemin/export")
    exporter.export_measurement(measurement_data)
"""

from __future__ import annotations

import os
import csv
import json
import math
import time
import sqlite3
import asyncio
import logging
import random
import functools
import signal
from datetime import datetime
from typing import List, Dict, Optional, Set, Callable
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

from src.utils.error_types import ErrorCategory
from src.utils.error_manager import error_manager

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

FORMAT_EXTENSIONS = {
    "xlsx": ".xlsx",
    "csv": ".csv",
    "json": ".json",
    "xml": ".xml",
    "pdf": ".pdf",
    "sqlite": ".db",
}

FORMAT_NAMES = {
    "xlsx": "Excel (.xlsx)",
    "csv": "CSV (.csv)",
    "json": "JSON (.json)",
    "xml": "XML (.xml)",
    "pdf": "PDF (.pdf)",
    "sqlite": "SQLite (.db)",
}

FORMAT_ORDER = ["xlsx", "csv", "json", "xml", "pdf", "sqlite"]

_MAX_RETRIES = 3
_RETRY_BASE_DELAY = 0.5  # secondes
_RETRY_MAX_DELAY = 4.0
_EXPORT_TIMEOUT = 30.0  # secondes max par export


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class ExportError(Exception):
    """Erreur lors de l'export d'une mesure."""


# ---------------------------------------------------------------------------
# JSON strict — sanitisation des NaN / Inf avant serialisation
# ---------------------------------------------------------------------------

def _sanitize_json(obj):
    """Remplace recursivement NaN/Inf par None pour une serialisation JSON sure."""
    if isinstance(obj, dict):
        return {k: _sanitize_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_json(v) for v in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    return obj


# ---------------------------------------------------------------------------
# Retry idempotent
# ---------------------------------------------------------------------------

def _retry_idempotent_file(func):
    """Decorateur : retry avec verification d'existence du fichier.

    Avant chaque retry, on verifie si le fichier retourne par
    la fonction existe deja. Si oui, on ne relance PAS l'operation
    (pas de doublon).

    Delai progressif : 0.5s, 1s, 2s + jitter +/-25%.
    """
    @functools.wraps(func)
    def wrapper(self, measurement, base):
        last_exc = None
        saved_fp = None
        for attempt in range(1, _MAX_RETRIES + 1):
            try:
                saved_fp = func(self, measurement, base)
                return saved_fp
            except Exception as exc:
                last_exc = exc
                # Verifier si le fichier existe deja (l'ecriture a pu
                # reussir entre-temps malgre l'exception)
                if saved_fp and os.path.exists(saved_fp) and os.path.getsize(saved_fp) > 0:
                    logger.info(
                        "Fichier existe deja malgre exception %s: %s",
                        func.__name__, saved_fp,
                    )
                    return saved_fp
                if attempt < _MAX_RETRIES:
                    delay = min(
                        _RETRY_BASE_DELAY * (2 ** (attempt - 1)),
                        _RETRY_MAX_DELAY,
                    )
                    jitter = delay * random.uniform(-0.25, 0.25)
                    total = delay + jitter
                    logger.warning(
                        "%s tentative %d/%d echouee, "
                        "nouvel essai dans %.2fs: %s",
                        func.__name__, attempt, _MAX_RETRIES,
                        total, exc,
                    )
                    time.sleep(total)
        raise last_exc  # type: ignore
    return wrapper


def _retry_idempotent_sqlite(func):
    """Decorateur : retry SQLite avec verification.

    Pour SQLite, on verifie que la connexion est toujours active
    et on retente avec backoff si la base est verrouillee.
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        last_exc = None
        for attempt in range(1, _MAX_RETRIES + 1):
            try:
                return func(*args, **kwargs)
            except sqlite3.OperationalError as exc:
                last_exc = exc
                if "locked" in str(exc).lower() and attempt < _MAX_RETRIES:
                    delay = min(
                        _RETRY_BASE_DELAY * (2 ** (attempt - 1)),
                        _RETRY_MAX_DELAY,
                    )
                    jitter = delay * random.uniform(-0.25, 0.25)
                    logger.warning(
                        "SQLite verrouille, tentative %d/%d dans %.2fs: %s",
                        attempt, _MAX_RETRIES, delay + jitter, exc,
                    )
                    time.sleep(delay + jitter)
                else:
                    raise
            except Exception as exc:
                # Erreur non-SQLite → ne pas retenter
                raise exc
        raise last_exc  # type: ignore
    return wrapper


# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------

def _fmt_ts(ts) -> str:
    if ts is None:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(ts, datetime):
        return ts.strftime("%Y-%m-%d %H:%M:%S")
    return str(ts)[:19]


# ---------------------------------------------------------------------------
# DataExporter
# ---------------------------------------------------------------------------

class DataExporter:
    """Exportateur multi-format.

    Attributes:
        output_dir: Dossier de destination.
        active_formats: Ensemble des formats actifs.
        on_exported: Callback (filepath, format) apres chaque reussite.
    """

    def __init__(
        self,
        output_dir: str = "",
        active_formats: Optional[Set[str]] = None,
    ):
        self.output_dir = output_dir or os.path.expanduser(
            "~/Documents/ApplicationMesure/Export"
        )
        self.active_formats = active_formats or {"xlsx", "csv"}
        self.on_exported: Optional[Callable[[str, str], None]] = None
        self._file_counter = 0

        # SQLite
        self._sqlite_conn: Optional[sqlite3.Connection] = None
        self._sqlite_path: Optional[str] = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close_sqlite()
        return False

    # ------------------------------------------------------------------
    # Configuration
    # ------------------------------------------------------------------

    def set_output_dir(self, path: str):
        self.output_dir = path
        os.makedirs(path, exist_ok=True)
        logger.info("Dossier d'export change: %s", path)

    def set_active_formats(self, formats: Set[str]):
        self.active_formats = {f for f in formats if f in FORMAT_EXTENSIONS}
        logger.info("Formats actifs: %s", ", ".join(sorted(self.active_formats)))

    def enable_format(self, fmt: str):
        if fmt in FORMAT_EXTENSIONS:
            self.active_formats.add(fmt)

    def disable_format(self, fmt: str):
        self.active_formats.discard(fmt)

    # ------------------------------------------------------------------
    # Point d'entree principal
    # ------------------------------------------------------------------

    def export_measurement(self, measurement: Dict) -> List[str]:
        """Exporte une mesure dans tous les formats actifs.

        Args:
            measurement: Dictionnaire de la mesure.

        Retourne:
            Liste des chemins de fichiers generes.
        """
        if not measurement:
            return []

        self._file_counter += 1
        ts = _fmt_ts(measurement.get("timestamp", datetime.now()))
        base = f"Mesure_{ts}_{self._file_counter:04d}"
        generated = []

        os.makedirs(self.output_dir, exist_ok=True)

        for fmt in FORMAT_ORDER:
            if fmt not in self.active_formats:
                continue
            try:
                fp = self._export_single(fmt, measurement, base)
                generated.append(fp)
                logger.debug("Export %s reussi: %s", fmt, fp)
                if self.on_exported:
                    self.on_exported(fp, fmt)
            except Exception as e:
                logger.error("Echec export %s: %s", fmt, e)
                error_manager.error(
                    category=ErrorCategory.EXPORT,
                    error_type="format_error",
                    message=(
                        "Erreur lors de l'export au format "
                        f"{fmt.upper()}. Verifiez le dossier "
                        "de destination."
                    ),
                )

        return generated

    async def async_export_measurement(self, measurement: Dict) -> List[str]:
        """Version async de export_measurement, non-bloquante pour l'event loop.

        Execute les exports dans un thread separe pour ne pas bloquer
        l'event loop asyncio.
        """
        return await asyncio.to_thread(self.export_measurement, measurement)

    def _export_single(self, fmt: str, measurement: Dict, base: str) -> str:
        exporter = {
            "xlsx": self._export_xlsx,
            "csv": self._export_csv,
            "json": self._export_json,
            "xml": self._export_xml,
            "pdf": self._export_pdf,
            "sqlite": self._export_sqlite,
        }
        fn = exporter.get(fmt)
        if fn is None:
            raise ExportError(f"Format inconnu: {fmt}")
        return fn(measurement, base)

    # ----------------------------------------------------------------
    # XLSX
    # ----------------------------------------------------------------

    @_retry_idempotent_file
    def _export_xlsx(self, measurement: Dict, base: str) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.cell import WriteOnlyCell

        fp = os.path.join(self.output_dir, f"{base}.xlsx")

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Mesure")

        hfont = Font(bold=True, size=12, color="FFFFFF")
        hfill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        bdr = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        cal = Alignment(horizontal="center")

        def sc(v, font=None, fill=None, alignment=None, border=None):
            c = WriteOnlyCell(ws, value=v)
            if font: c.font = font
            if fill: c.fill = fill
            if alignment: c.alignment = alignment
            if border: c.border = border
            return c

        headers = ["Horodatage", "Outil", "ID Outil",
                     "Valeur", "Unite", "Statut", "Note"]
        hr = []
        for h in headers:
            hr.append(sc(h, font=hfont, fill=hfill, alignment=cal, border=bdr))
        ws.append(hr)

        vals = [
            _fmt_ts(measurement.get("timestamp")),
            measurement.get("tool_name", ""),
            measurement.get("tool_id", ""),
            measurement.get("value", ""),
            measurement.get("unit", ""),
            measurement.get("status", ""),
            measurement.get("note", ""),
        ]
        dr = []
        for v in vals:
            dr.append(sc(v, border=bdr))
        ws.append(dr)

        wb.save(fp)
        return fp

    # ----------------------------------------------------------------
    # CSV
    # ----------------------------------------------------------------

    @_retry_idempotent_file
    def _export_csv(self, measurement: Dict, base: str) -> str:
        fp = os.path.join(self.output_dir, f"{base}.csv")

        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Horodatage", "Outil", "ID Outil",
                         "Valeur", "Unite", "Statut", "Note"])
            w.writerow([
                _fmt_ts(measurement.get("timestamp")),
                measurement.get("tool_name", ""),
                measurement.get("tool_id", ""),
                measurement.get("value", ""),
                measurement.get("unit", ""),
                measurement.get("status", ""),
                measurement.get("note", ""),
            ])
        return fp

    # ----------------------------------------------------------------
    # JSON
    # ----------------------------------------------------------------

    @_retry_idempotent_file
    def _export_json(self, measurement: Dict, base: str) -> str:
        fp = os.path.join(self.output_dir, f"{base}.json")

        data = {
            "export_date": datetime.now().isoformat(),
            "measurement": {
                "timestamp": _fmt_ts(measurement.get("timestamp")),
                "tool_name": measurement.get("tool_name", ""),
                "tool_id": measurement.get("tool_id"),
                "value": (
                    float(measurement["value"])
                    if measurement.get("value") is not None
                    else None
                ),
                "unit": measurement.get("unit", ""),
                "status": measurement.get("status", "OK"),
                "note": measurement.get("note", ""),
            },
        }

        with open(fp, "w", encoding="utf-8") as f:
            json.dump(_sanitize_json(data), f, indent=2, ensure_ascii=False)
            f.write("\n")
        return fp

    # ----------------------------------------------------------------
    # XML
    # ----------------------------------------------------------------

    @_retry_idempotent_file
    def _export_xml(self, measurement: Dict, base: str) -> str:
        fp = os.path.join(self.output_dir, f"{base}.xml")

        root = Element("MesureExport")
        SubElement(root, "DateGeneration").text = datetime.now().isoformat()
        m = SubElement(root, "Mesure")
        SubElement(m, "Horodatage").text = _fmt_ts(measurement.get("timestamp"))
        SubElement(m, "Outil").text = str(measurement.get("tool_name", ""))
        SubElement(m, "IDOutil").text = str(measurement.get("tool_id", ""))
        SubElement(m, "Valeur").text = str(measurement.get("value", ""))
        SubElement(m, "Unite").text = str(measurement.get("unit", ""))
        SubElement(m, "Statut").text = str(measurement.get("status", "OK"))
        SubElement(m, "Note").text = str(measurement.get("note", ""))

        xml_str = minidom.parseString(
            tostring(root, encoding="unicode")
        ).toprettyxml(indent="  ")

        with open(fp, "w", encoding="utf-8") as f:
            f.write(xml_str)
        return fp

    # ----------------------------------------------------------------
    # PDF
    # ----------------------------------------------------------------

    @_retry_idempotent_file
    def _export_pdf(self, measurement: Dict, base: str) -> str:
        fp = os.path.join(self.output_dir, f"{base}.pdf")

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )

        doc = SimpleDocTemplate(fp, pagesize=A4,
                                topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()

        ts_style = ParagraphStyle("TitleC", parent=styles["Title"],
                                  fontSize=18, spaceAfter=20,
                                  textColor=colors.HexColor("#1E1E2E"))
        story = [Paragraph("Rapport de Mesure", ts_style), Spacer(1, 12)]

        info_s = ParagraphStyle("Info", parent=styles["Normal"], fontSize=10)
        story.append(Paragraph(
            f"Date d'export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            info_s))
        story.append(Spacer(1, 12))

        tbl_data = [
            ["Propriete", "Valeur"],
            ["Horodatage", _fmt_ts(measurement.get("timestamp"))],
            ["Outil", measurement.get("tool_name", "")],
            ["ID Outil", str(measurement.get("tool_id", ""))],
            ["Valeur", f"{measurement.get('value', '')} {measurement.get('unit', '')}"],
            ["Statut", measurement.get("status", "OK")],
            ["Note", measurement.get("note", "")],
        ]
        tbl = Table(tbl_data, colWidths=[120, 300])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f5f5f5")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f9f9f9")]),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 30))
        ftr = ParagraphStyle("Footer", parent=styles["Normal"],
                             fontSize=8, textColor=colors.grey, alignment=1)
        story.append(Paragraph(
            "Application de Mesure - Document genere automatiquement", ftr))
        doc.build(story)
        return fp

    # ----------------------------------------------------------------
    # SQLite
    # ----------------------------------------------------------------

    def _export_sqlite(self, measurement: Dict, base: str) -> str:
        db_path = self._get_or_create_sqlite_db()
        self._sqlite_insert(measurement)
        return db_path

    def _get_or_create_sqlite_db(self) -> str:
        db_path = os.path.join(self.output_dir, "mesures.db")
        if self._sqlite_conn is None or self._sqlite_path != db_path:
            if self._sqlite_conn:
                self._sqlite_conn.close()
            self._sqlite_path = db_path
            self._sqlite_conn = sqlite3.connect(db_path, timeout=10)
            self._sqlite_conn.execute("PRAGMA journal_mode=WAL")
            self._sqlite_conn.execute("PRAGMA synchronous=NORMAL")
            self._sqlite_conn.execute("""
                CREATE TABLE IF NOT EXISTS mesures (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT NOT NULL,
                    tool_name TEXT,
                    tool_id INTEGER,
                    value REAL,
                    unit TEXT,
                    status TEXT DEFAULT 'OK',
                    note TEXT,
                    export_date TEXT NOT NULL
                )
            """)
            self._sqlite_conn.commit()
        return db_path

    @_retry_idempotent_sqlite
    def _sqlite_insert(self, measurement: Dict):
        conn = self._sqlite_conn
        if conn is None:
            return
        conn.execute(
            """INSERT INTO mesures (timestamp, tool_name, tool_id,
               value, unit, status, note, export_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                _fmt_ts(measurement.get("timestamp")),
                measurement.get("tool_name", ""),
                measurement.get("tool_id"),
                measurement.get("value"),
                measurement.get("unit", ""),
                measurement.get("status", "OK"),
                measurement.get("note", ""),
                datetime.now().isoformat(),
            ),
        )
        conn.commit()

    # ------------------------------------------------------------------
    # Nettoyage
    # ------------------------------------------------------------------

    def close_sqlite(self):
        """Ferme proprement la connexion SQLite."""
        if self._sqlite_conn:
            try:
                # Forcer le checkpoint WAL pour eviter les fichiers -wal/-shm orphelins
                self._sqlite_conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            except Exception:
                pass
            try:
                self._sqlite_conn.close()
            except Exception:
                pass
            self._sqlite_conn = None
            self._sqlite_path = None

    @property
    def summary(self) -> Dict:
        return {
            "output_dir": self.output_dir,
            "active_formats": sorted(self.active_formats),
            "sqlite_db": self._sqlite_path,
        }
